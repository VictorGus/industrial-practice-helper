import asyncio
import io
import re
import zipfile

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, Update
from telegram.ext import ContextTypes

from common.config import get_group_number_regex, get_webdav_upload_dir
from common.logger import log
from common.storage import (
    upload_zip, upload_single_member_zip, upload_bytes,
    list_group_xlsx_files, list_group_zip_files, download_xlsx, download_file,
    sync_group_xlsx,
    list_student_folders, list_students, set_student_comment,
    is_admin, REQUIRED_DOCUMENTS, _fix_zip_filename,
)

# Russian name: capital letter followed by lowercase Cyrillic
_RU_NAME = r"[А-ЯЁ][а-яё]+"
_FOLDER_RE = re.compile(rf"^({_RU_NAME})_({_RU_NAME})(_({_RU_NAME})?)?$")


def _build_group_zip_regex() -> re.Pattern:
    group = get_group_number_regex()
    return re.compile(rf"^({group})\.zip$")


def _build_member_zip_regex() -> re.Pattern:
    group = get_group_number_regex()
    return re.compile(rf"^({group})_({_RU_NAME})_({_RU_NAME})(_({_RU_NAME})?)?\.zip$")


def _get_top_level_folders(zf: zipfile.ZipFile) -> list[str]:
    folders = set()
    for entry in zf.infolist():
        parts = _fix_zip_filename(entry).split("/")
        if len(parts) > 1 and parts[0]:
            folders.add(parts[0])
    return sorted(folders)


def _get_student_folders(zf: zipfile.ZipFile, group_number: str) -> tuple[list[str], bool]:
    top_folders = _get_top_level_folders(zf)

    if len(top_folders) == 1 and top_folders[0] == group_number:
        folders = set()
        for entry in zf.infolist():
            parts = _fix_zip_filename(entry).split("/")
            if len(parts) > 2 and parts[0] == group_number and parts[1]:
                folders.add(parts[1])
        return sorted(folders), True

    return top_folders, False


def _validate_group_zip(filename: str, group_number: str, buf: io.BytesIO) -> tuple[bool, bool, list[str]]:
    errors = []
    buf.seek(0)
    try:
        with zipfile.ZipFile(buf) as zf:
            folders, is_wrapped = _get_student_folders(zf, group_number)
            if not folders:
                errors.append("Архив не содержит папок студентов.")
                return False, False, errors

            invalid_folders = [f for f in folders if not _FOLDER_RE.match(f)]
            if invalid_folders:
                errors.append("Следующие папки имеют некорректные имена:")
                for f in invalid_folders:
                    errors.append(f"  • {f}")
                errors.append("")
                errors.append(
                    "Имя папки должно соответствовать формату:\n"
                    "  {Фамилия}_{Имя}_{Отчество}\n"
                    "Пример: Иванов_Иван_Иванович\n"
                    "Отчество можно опустить:\n"
                    "  Иванов_Иван_"
                )
                return False, is_wrapped, errors
    except zipfile.BadZipFile:
        errors.append("Файл не является корректным ZIP-архивом.")
        return False, False, errors

    return True, is_wrapped, []


def _parse_filename(filename: str) -> tuple[str, str | None]:
    m = _build_group_zip_regex().match(filename)
    if m:
        return "group", m.group(1)

    m = _build_member_zip_regex().match(filename)
    if m:
        return "member", m.group(1)

    return "invalid", None


def _member_folder_name(filename: str) -> str:
    name = filename[:-4]
    _, rest = name.split("_", 1)
    return rest


import time


def _make_sync_progress_callback(message, loop):
    """Progress callback for sync: shows student name being processed."""
    last_edit = [0.0]

    def on_progress(done, total, student_name):
        now = time.monotonic()
        if done < total and now - last_edit[0] < 3.0:
            return
        last_edit[0] = now
        bar_len = 10
        filled = int(bar_len * done / total) if total else bar_len
        bar = "▓" * filled + "░" * (bar_len - filled)
        text = f"🔄 Синхронизация: {bar} {done}/{total}\n{student_name}"
        asyncio.run_coroutine_threadsafe(
            message.edit_text(text), loop
        )

    return on_progress


def _make_progress_callback(message, loop):
    """Create a throttled progress callback that edits a Telegram message.

    Called from a worker thread; schedules edits on the event loop.
    """
    last_edit = [0.0]

    def on_progress(done, total):
        now = time.monotonic()
        # Throttle: update at most every 3 seconds, or on completion
        if done < total and now - last_edit[0] < 3.0:
            return
        last_edit[0] = now
        bar_len = 10
        filled = int(bar_len * done / total) if total else bar_len
        bar = "▓" * filled + "░" * (bar_len - filled)
        text = f"⏳ Загрузка: {bar} {done}/{total}"
        asyncio.run_coroutine_threadsafe(
            message.edit_text(text), loop
        )

    return on_progress


def _display_group(group: str) -> str:
    """Convert group name for display: 3130801-30201 -> 3130801/30201."""
    return group.replace("-", "/")


def _user_info(update: Update) -> str:
    user = update.effective_user
    if user:
        return f"@{user.username}" if user.username else f"{user.full_name} (id={user.id})"
    return "unknown"


_DENY_MSG = "⛔ У вас нет доступа. Обратитесь к администратору."


async def _check_admin(update: Update) -> bool:
    user = update.effective_user
    username = user.username if user else None
    if is_admin(username):
        return True
    log.warning("Access denied for %s", _user_info(update))
    target = update.message or (update.callback_query and update.callback_query.message)
    if target:
        await target.reply_text(_DENY_MSG)
    return False


BTN_HELP = "❓ Помощь"
BTN_STATUS = "📊 Статус"
BTN_SYNC = "🔄 Синхронизация"
BTN_COMMENT = "💬 Комментарий"
BTN_UPLOAD_XLSX = "📤 Синхронизация учебных групп"
BTN_UNPACK_ZIP = "📦 Распаковать архив"

_ADMIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        [BTN_STATUS, BTN_SYNC],
        [BTN_COMMENT, BTN_UPLOAD_XLSX],
        [BTN_UNPACK_ZIP, BTN_HELP],
    ],
    resize_keyboard=True,
)

_PUBLIC_KEYBOARD = ReplyKeyboardMarkup(
    [[BTN_HELP]],
    resize_keyboard=True,
)


_HELP_TEXT = r"""
📋 Бот для сбора документов для прохождения практики.

📎 Как отправить документы:

Для одного студента — отправьте .zip архив с именем:
  {Группа}\_{Фамилия}\_{Имя}\_{Отчество}.zip
  Пример: 3130801-30201\_Иванов\_Иван\_Иванович.zip

Для всей группы — отправьте .zip архив с именем:
  {Группа}.zip
  Пример: 3130801-30201.zip

📁 Структура архива группы:
  Каждый студент — отдельная папка:
    Иванов\_Иван\_Иванович/
    Петрова\_Мария\_Сергеевна/

📄 Необходимые документы (в папке каждого студента):

❗ Имя файла должно *содержать* название документа:
  • Гарантийное письмо
  • Заявление
  • Краткосрочный договор
  • Характеристика
Формат файла — любой (docx, pdf, doc и т.д.)

⚠️ В имени файла "/" в номере группы заменяется на "-"
  Группа 3130801/30201 → файл 3130801-30201.zip

ℹ️ Отчество можно опустить: Иванов\_Иван\_
"""


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    log.info("User %s sent /help", _user_info(update))
    await update.message.reply_text(_HELP_TEXT, parse_mode="Markdown")


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    log.info("User %s sent /start", _user_info(update))
    user = update.effective_user
    username = user.username if user else None
    keyboard = _ADMIN_KEYBOARD if is_admin(username) else _PUBLIC_KEYBOARD
    await update.message.reply_text(
        "Привет! Я бот для сбора документов для прохождения практики",
        reply_markup=keyboard,
    )
    await update.message.reply_text(_HELP_TEXT, parse_mode="Markdown")


def _parse_xlsx(buf: io.BytesIO) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active

    # Map column headers to indices
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    doc_columns = {}
    for col_header in REQUIRED_DOCUMENTS:
        if col_header in headers:
            doc_columns[col_header] = headers.index(col_header)
    comment_idx = headers.index("Комментарий") if "Комментарий" in headers else None

    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        missing = []
        for col_header, idx in doc_columns.items():
            val = str(row[idx] or "").strip().lower()
            if val not in ("да", "yes", "1", "true", "не нужно", "не надо"):
                missing.append(col_header)
        comment = ""
        if comment_idx is not None and row[comment_idx]:
            comment = str(row[comment_idx]).strip()
        students.append({
            "surname": row[1] or "",
            "name": row[2] or "",
            "patronymic": row[3] or "",
            "missing": missing,
            "comment": comment,
        })
    wb.close()
    return students


async def _status_for_group(group: str) -> str:
    buf = await asyncio.to_thread(download_xlsx, group)
    students = _parse_xlsx(buf)
    incomplete = [s for s in students if s["missing"]]
    total = len(students)

    dg = _display_group(group)
    lines = [f"📁 Группа {dg} — сдали всё: {total - len(incomplete)}/{total}", ""]
    for s in students:
        full_name = " ".join(filter(None, [s["surname"], s["name"], s["patronymic"]]))
        if not s["missing"]:
            lines.append(f"✅ {full_name}")
        else:
            docs = ", ".join(s["missing"])
            lines.append(f"❌ {full_name}")
            lines.append(f"     нет: {docs}")
        if s["comment"]:
            lines.append(f"     💬 {s['comment']}")
    return "\n".join(lines)


async def _group_picker(update: Update, action: str, label: str) -> None:
    """Show inline buttons to pick a group, plus an 'All' option for status."""
    try:
        groups = await asyncio.to_thread(list_group_xlsx_files)
    except Exception as e:
        log.error("Failed to list groups: %s", e)
        await update.message.reply_text(f"Не удалось получить список групп: {e}")
        return

    if not groups:
        await update.message.reply_text("Файлы групп (.xlsx) не найдены.")
        return

    buttons = [[InlineKeyboardButton(_display_group(g), callback_data=f"{action}:{g}")] for g in groups]
    if action == "status":
        buttons.append([InlineKeyboardButton("Все группы", callback_data="status:__all__")])

    await update.message.reply_text(
        f"{label} — выберите группу:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    log.info("User %s sent /status", _user_info(update))
    if not await _check_admin(update):
        return
    if context.args:
        group = context.args[0]
        try:
            text = await _status_for_group(group)
        except Exception as e:
            log.error("Failed to read %s.xlsx: %s", group, e)
            text = f"Не удалось прочитать {_display_group(group)}.xlsx: {e}"
        await update.message.reply_text(text)
        return
    await _group_picker(update, "status", "📊 Статус")


async def sync(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    log.info("User %s sent /sync", _user_info(update))
    if not await _check_admin(update):
        return
    if context.args:
        group = context.args[0]
        await _run_sync(update.message, group)
        return
    await _group_picker(update, "sync", "🔄 Синхронизация")


async def _run_sync(message, group: str) -> None:
    """Run sync with progress bar, used by both command and callback."""
    dg = _display_group(group)
    progress_msg = await message.reply_text(f"🔄 Синхронизация {dg}: ░░░░░░░░░░ 0/?")
    loop = asyncio.get_running_loop()
    on_progress = _make_sync_progress_callback(progress_msg, loop)
    try:
        updated = await asyncio.to_thread(sync_group_xlsx, group, on_progress)
    except Exception as e:
        log.error("Failed to sync %s: %s", group, e)
        await progress_msg.edit_text(f"❌ Не удалось синхронизировать {dg}: {e}")
        return
    if updated:
        await progress_msg.edit_text(f"✅ Синхронизация {dg}: обновлено {updated} ячеек.")
    else:
        await progress_msg.edit_text(f"✅ Синхронизация {dg}: xlsx уже актуален.")


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    query = update.callback_query
    await query.answer()
    if not await _check_admin(update):
        return
    data = query.data
    log.info("Callback from %s: %s", _user_info(update), data)

    if data.startswith("status:"):
        group = data.split(":", 1)[1]
        if group == "__all__":
            try:
                groups = await asyncio.to_thread(list_group_xlsx_files)
            except Exception as e:
                await query.message.reply_text(f"Ошибка: {e}")
                return
            for g in groups:
                try:
                    text = await _status_for_group(g)
                except Exception as e:
                    text = f"📁 Группа {_display_group(g)}: ошибка чтения файла"
                await query.message.reply_text(text)
        else:
            try:
                text = await _status_for_group(group)
            except Exception as e:
                text = f"Не удалось прочитать {_display_group(group)}.xlsx: {e}"
            await query.message.reply_text(text)

    elif data.startswith("sync:"):
        group = data.split(":", 1)[1]
        await _run_sync(query.message, group)

    elif data.startswith("comment_group:"):
        group = data.split(":", 1)[1]
        try:
            students = await asyncio.to_thread(list_students, group)
        except Exception as e:
            await query.message.reply_text(f"Ошибка: {e}")
            return
        if not students:
            await query.message.reply_text("Список студентов пуст.")
            return
        # Store student list in user_data, use index in callback
        context.user_data["comment_students"] = {"group": group, "list": students}
        buttons = []
        for i, s in enumerate(students):
            full_name = " ".join(filter(None, [s["surname"], s["name"], s["patronymic"]]))
            buttons.append([InlineKeyboardButton(full_name, callback_data=f"cs:{i}")])
        await query.message.reply_text(
            f"Группа {_display_group(group)} — выберите студента:",
            reply_markup=InlineKeyboardMarkup(buttons),
        )

    elif data.startswith("unpack:"):
        zip_filename = data.split(":", 1)[1]
        group_number = zip_filename[:-4]  # strip .zip
        dg = _display_group(group_number)

        progress_msg = await query.message.reply_text(
            f"📦 Скачиваем архив {zip_filename}…"
        )

        try:
            upload_dir = get_webdav_upload_dir()
            remote_zip = f"{upload_dir}{zip_filename}"
            buf = await asyncio.to_thread(download_file, remote_zip)
        except Exception as e:
            log.error("Failed to download %s: %s", zip_filename, e)
            await progress_msg.edit_text(
                f"❌ Не удалось скачать архив {zip_filename}: {e}"
            )
            return

        # Validate zip
        valid, is_wrapped, errors = _validate_group_zip(zip_filename, group_number, buf)
        if not valid:
            lines = [f"❌ Архив {zip_filename} не прошёл валидацию:", ""]
            lines.extend(errors)
            await progress_msg.edit_text("\n".join(lines))
            return

        # Unpack and merge into group directory
        try:
            loop = asyncio.get_running_loop()
            on_progress = _make_progress_callback(progress_msg, loop)
            remote_paths = await asyncio.to_thread(
                upload_zip, buf, zip_filename, is_wrapped, on_progress,
                True,  # skip_zip_upload — archive is already on disk
            )
            log.info("Unpacked %s into group dir (%d files)", zip_filename, len(remote_paths))
        except Exception as e:
            log.error("Failed to unpack %s: %s", zip_filename, e)
            await progress_msg.edit_text(
                f"❌ Ошибка распаковки {zip_filename}: {e}"
            )
            return

        # Sync xlsx
        try:
            synced = await asyncio.to_thread(sync_group_xlsx, group_number)
            if synced:
                log.info("Synced %d cell(s) in %s.xlsx after unpack", synced, group_number)
        except Exception as e:
            log.warning("Failed to sync %s.xlsx: %s", group_number, e)

        await progress_msg.edit_text(
            f"✅ Архив {zip_filename} распакован!\n"
            f"Группа {dg}: загружено {len(remote_paths)} файл(ов)."
        )

    elif data.startswith("cs:"):
        idx = int(data.split(":", 1)[1])
        stored = context.user_data.get("comment_students")
        if not stored or idx >= len(stored["list"]):
            await query.message.reply_text("Ошибка. Попробуйте заново.")
            return
        group = stored["group"]
        s = stored["list"][idx]
        full_name = " ".join(filter(None, [s["surname"], s["name"], s["patronymic"]]))
        context.user_data["comment_for"] = {
            "group": group,
            "surname": s["surname"],
            "name": s["name"],
            "patronymic": s["patronymic"],
        }
        del context.user_data["comment_students"]
        await query.message.reply_text(f"Введите комментарий для {full_name} (группа {_display_group(group)}):")


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    text = update.message.text
    log.info("User %s sent message: %s", _user_info(update), text)

    if text == BTN_HELP:
        await update.message.reply_text(_HELP_TEXT, parse_mode="Markdown")
        return

    if not await _check_admin(update):
        return

    # Check if user is entering a comment
    comment_for = context.user_data.get("comment_for")
    if comment_for and text not in (BTN_STATUS, BTN_SYNC, BTN_COMMENT, BTN_UNPACK_ZIP):
        del context.user_data["comment_for"]
        group = comment_for["group"]
        surname = comment_for["surname"]
        name = comment_for["name"]
        patronymic = comment_for["patronymic"]
        full_name = " ".join(filter(None, [surname, name, patronymic]))
        try:
            ok = await asyncio.to_thread(
                set_student_comment, group, surname, name, patronymic, text,
            )
        except Exception as e:
            await update.message.reply_text(f"Ошибка: {e}")
            return
        if ok:
            await update.message.reply_text(f"💬 Комментарий для {full_name} сохранён.")
        else:
            await update.message.reply_text(
                f"Не удалось сохранить комментарий. Проверьте, что {_display_group(group)}.xlsx "
                "содержит колонку «Комментарий» и студент существует."
            )
        return

    # Any button press cancels awaiting states
    context.user_data.pop("awaiting_xlsx", None)

    if text == BTN_STATUS:
        context.args = []
        await status(update, context)
    elif text == BTN_SYNC:
        context.args = []
        await sync(update, context)
    elif text == BTN_COMMENT:
        context.user_data.pop("comment_for", None)
        await _group_picker(update, "comment_group", "💬 Комментарий")
    elif text == BTN_UPLOAD_XLSX:
        context.user_data["awaiting_xlsx"] = True
        await update.message.reply_text(
            "Отправьте один или несколько .xlsx файлов групп.\n"
            "Файлы будут загружены в /Практики/ на Яндекс Диск.\n"
            "Для завершения нажмите любую другую кнопку."
        )
    elif text == BTN_UNPACK_ZIP:
        try:
            zips = await asyncio.to_thread(list_group_zip_files)
        except Exception as e:
            log.error("Failed to list group zips: %s", e)
            await update.message.reply_text(f"Не удалось получить список архивов: {e}")
            return
        if not zips:
            await update.message.reply_text("Групповые архивы (.zip) не найдены на Яндекс Диске.")
            return
        buttons = [
            [InlineKeyboardButton(z, callback_data=f"unpack:{z}")]
            for z in zips
        ]
        await update.message.reply_text(
            "📦 Распаковка архива — выберите архив:",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
    else:
        await update.message.reply_text(text)


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    doc = update.message.document
    if not doc:
        return

    user = _user_info(update)
    filename = doc.file_name or ""

    # Handle xlsx upload (flag stays active until user presses another button)
    if context.user_data.get("awaiting_xlsx"):
        if not await _check_admin(update):
            return
        if not filename.lower().endswith(".xlsx"):
            await update.message.reply_text("Пожалуйста, отправьте файл в формате .xlsx.")
            return
        log.info("User %s uploading xlsx: %s", user, filename)
        file = await doc.get_file()
        buf = io.BytesIO()
        await file.download_to_memory(buf)
        try:
            upload_dir = get_webdav_upload_dir()
            remote_path = f"{upload_dir}{filename}"
            await asyncio.to_thread(upload_bytes, buf, remote_path)
            log.info("Uploaded %s to %s", filename, remote_path)
            await update.message.reply_text(f"✅ {filename} загружен на Яндекс Диск.")
        except Exception as e:
            log.error("Failed to upload %s: %s", filename, e)
            await update.message.reply_text(f"❌ {filename}: ошибка загрузки — {e}")
        return

    if not filename.lower().endswith(".zip"):
        log.warning("User %s sent non-zip file: %s", user, filename)
        await update.message.reply_text("Пожалуйста, отправьте файл в формате .zip.")
        return

    log.info("User %s uploaded file: %s (%d bytes)", user, filename, doc.file_size)

    # Download
    file = await doc.get_file()
    buf = io.BytesIO()
    await file.download_to_memory(buf)

    # Log zip contents
    buf.seek(0)
    try:
        with zipfile.ZipFile(buf) as zf:
            entries = zf.infolist()
            total_size = sum(e.file_size for e in entries)
            log.info("ZIP %s: %d items, %d bytes uncompressed", filename, len(entries), total_size)
            for e in entries:
                log.info("  %s (%d bytes)", _fix_zip_filename(e), e.file_size)
    except zipfile.BadZipFile:
        log.error("Invalid ZIP file: %s", filename)

    # Determine type
    file_type, group_number = _parse_filename(filename)

    if file_type == "invalid":
        log.warning("Validation failed for %s from %s: filename doesn't match any pattern", filename, user)

        example_group = "3130801-30201"
        status_msg = await update.message.reply_text(
            "Ошибка валидации:\n\n"
            f"Имя архива «{filename}» не соответствует ожидаемому формату:\n"
            f"  • {{Группа}}.zip (напр. {example_group}.zip) — для всей группы\n"
            f"  • {{Группа}}_{{Фамилия}}_{{Имя}}_{{Отчество}}.zip "
            f"(напр. {example_group}_Иванов_Иван_Иванович.zip) — для одного студента\n\n"
            "⏳ Сохраняем файл на Яндекс Диск…\n"
            "Рекомендуем переименовать файл и загрузить заново."
        )

        # Save the raw .zip to Yandex Disk (no extraction)
        try:
            upload_dir = get_webdav_upload_dir()
            remote_path = f"{upload_dir}{filename}"
            await asyncio.to_thread(upload_bytes, buf, remote_path)
            log.info("Saved raw zip %s to %s (validation failed)", filename, remote_path)
            await status_msg.edit_text(
                "Ошибка валидации:\n\n"
                f"Имя архива «{filename}» не соответствует ожидаемому формату:\n"
                f"  • {{Группа}}.zip (напр. {example_group}.zip) — для всей группы\n"
                f"  • {{Группа}}_{{Фамилия}}_{{Имя}}_{{Отчество}}.zip "
                f"(напр. {example_group}_Иванов_Иван_Иванович.zip) — для одного студента\n\n"
                "✅ Файл сохранён на Яндекс Диск, но его содержимое не будет учтено автоматически.\n"
                "Рекомендуем переименовать файл и загрузить заново."
            )
        except Exception as e:
            log.error("Failed to save raw zip %s: %s", filename, e)
            await status_msg.edit_text(
                "Ошибка валидации:\n\n"
                f"Имя архива «{filename}» не соответствует ожидаемому формату:\n"
                f"  • {{Группа}}.zip (напр. {example_group}.zip) — для всей группы\n"
                f"  • {{Группа}}_{{Фамилия}}_{{Имя}}_{{Отчество}}.zip "
                f"(напр. {example_group}_Иванов_Иван_Иванович.zip) — для одного студента\n\n"
                "❌ Не удалось сохранить файл на Яндекс Диск.\n"
                "Рекомендуем переименовать файл и загрузить заново."
            )


        return

    if file_type == "group":
        valid, is_wrapped, errors = _validate_group_zip(filename, group_number, buf)
        if not valid:
            log.warning("Validation failed for %s from %s: %s", filename, user, "; ".join(errors))
            lines = ["Ошибка валидации:", ""]
            lines.extend(errors)
            lines.append("")
            lines.append("Исправьте ошибки и загрузите архив заново.")
            await update.message.reply_text("\n".join(lines))
            return

        try:
            progress_msg = await update.message.reply_text("⏳ Загрузка: ░░░░░░░░░░ 0/?")
            loop = asyncio.get_running_loop()
            on_progress = _make_progress_callback(progress_msg, loop)

            remote_paths = await asyncio.to_thread(
                upload_zip, buf, filename, is_wrapped, on_progress,
            )
            log.info("Group archive %s from %s uploaded (%d files)", filename, user, len(remote_paths))
            for rp in remote_paths:
                log.info("  -> %s", rp)

            # Sync xlsx with actual files on disk
            try:
                synced = await asyncio.to_thread(sync_group_xlsx, group_number)
                if synced:
                    log.info("Synced %d cell(s) in %s.xlsx", synced, group_number)
            except Exception as e:
                log.warning("Failed to sync %s.xlsx: %s", group_number, e)

            lines = [
                f"✅ Архив группы {filename} принят!",
                f"Загружено {len(remote_paths)} файл(ов) на Яндекс Диск.",
            ]
            await progress_msg.edit_text("\n".join(lines))
        except Exception as e:
            log.error("Upload failed for %s from %s: %s", filename, user, e)
            await update.message.reply_text(f"Ошибка загрузки на Яндекс Диск: {e}")

    elif file_type == "member":
        folder_name = _member_folder_name(filename)
        try:
            progress_msg = await update.message.reply_text("⏳ Загрузка: ░░░░░░░░░░ 0/?")
            loop = asyncio.get_running_loop()
            on_progress = _make_progress_callback(progress_msg, loop)

            remote_paths = await asyncio.to_thread(
                upload_single_member_zip, buf, group_number, folder_name, on_progress,
            )
            log.info("Student archive %s from %s uploaded to group %s as %s/ (%d files)",
                     filename, user, group_number, folder_name, len(remote_paths))
            for rp in remote_paths:
                log.info("  -> %s", rp)

            # Sync xlsx with actual files on disk
            try:
                synced = await asyncio.to_thread(sync_group_xlsx, group_number)
                if synced:
                    log.info("Synced %d cell(s) in %s.xlsx", synced, group_number)
            except Exception as e:
                log.warning("Failed to sync %s.xlsx: %s", group_number, e)

            lines = [
                f"✅ Архив студента {filename} принят!",
                f"Загружено в группу {_display_group(group_number)} как {folder_name}/",
                f"{len(remote_paths)} файл(ов) загружено на Яндекс Диск.",
            ]
            await progress_msg.edit_text("\n".join(lines))
        except Exception as e:
            log.error("Upload failed for %s from %s: %s", filename, user, e)
            await update.message.reply_text(f"Ошибка загрузки на Яндекс Диск: {e}")
