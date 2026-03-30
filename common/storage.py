import io
import re
import zipfile

from webdav3.client import Client

from common.config import get_webdav_options, get_webdav_upload_dir
from common.logger import log


def _fix_zip_filename(entry: zipfile.ZipInfo) -> str:
    """Fix Cyrillic filenames in ZIP archives created on Russian Windows.

    Python's zipfile decodes non-UTF-8 entries as CP437, but Russian Windows
    tools typically encode filenames in CP866.  Re-encode from CP437 back to
    bytes and decode as CP866 when the UTF-8 flag is not set.
    """
    if entry.flag_bits & 0x800:  # UTF-8 flag is set — already correct
        return entry.filename
    try:
        return entry.filename.encode("cp437").decode("cp866")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return entry.filename


SUPER_ADMINS = {"victorgus", "anizhomlev", "polly1821"}


def _get_client() -> Client:
    return Client(get_webdav_options())


def get_admin_usernames() -> set[str]:
    """Read admin usernames from Администраторы.xlsx on Yandex Disk.

    Returns a set of usernames (without @) in lowercase.
    """
    from openpyxl import load_workbook

    client = _get_client()
    upload_dir = get_webdav_upload_dir()
    remote_path = f"{upload_dir}Администраторы.xlsx"

    if not client.check(remote_path):
        return set()

    buf = io.BytesIO()
    client.download_from(buf, remote_path)
    buf.seek(0)

    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    usernames = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            name = str(row[0]).strip().lstrip("@").lower()
            if name:
                usernames.add(name)
    wb.close()
    return usernames


def is_admin(username: str | None) -> bool:
    """Check if a Telegram username is an admin."""
    if not username:
        return False
    if username.lower() in SUPER_ADMINS:
        return True
    return username.lower() in get_admin_usernames()


def ensure_remote_dir(path: str) -> None:
    client = _get_client()
    if not client.check(path):
        client.mkdir(path)


def upload_bytes(buf: io.BytesIO, remote_path: str) -> None:
    client = _get_client()
    buf.seek(0)
    client.upload_to(buf, remote_path)


def _count_zip_files(buf: io.BytesIO, strip_prefix: str = "") -> int:
    """Count non-directory entries in a zip, optionally filtering by prefix."""
    buf.seek(0)
    with zipfile.ZipFile(buf) as zf:
        count = 0
        for entry in zf.infolist():
            if entry.is_dir():
                continue
            path = _fix_zip_filename(entry)
            if strip_prefix and not path.startswith(strip_prefix):
                continue
            rel = path[len(strip_prefix):]
            if rel:
                count += 1
    return count


def upload_zip(buf: io.BytesIO, filename: str, is_wrapped: bool = False,
               on_progress=None) -> list[str]:
    """Upload .zip and extract contents into a group folder.

    Zip name is {GROUP}.zip. Contents are extracted into
    {UPLOAD_DIR}/{GROUP}/ — the group folder is created if it doesn't exist,
    or accumulated into if it does.

    If is_wrapped is True, the archive has a top-level {GROUP}/ folder that
    should be stripped (its contents go directly into the group dir).

    on_progress(uploaded_count, total_count) is called after each file upload.

    Returns list of uploaded remote paths.
    """
    upload_dir = get_webdav_upload_dir()
    ensure_remote_dir(upload_dir)

    uploaded = []

    # Upload the .zip itself
    zip_remote = f"{upload_dir}{filename}"
    upload_bytes(buf, zip_remote)
    uploaded.append(zip_remote)

    # Group folder: {UPLOAD_DIR}/{GROUP}/
    group_name = filename[:-4]  # strip .zip
    group_dir = f"{upload_dir}{group_name}/"
    ensure_remote_dir(group_dir)

    # Prefix to strip from entry paths when wrapped
    strip_prefix = f"{group_name}/" if is_wrapped else ""

    total = _count_zip_files(buf, strip_prefix) + 1  # +1 for the zip itself
    if on_progress:
        on_progress(1, total)

    buf.seek(0)
    with zipfile.ZipFile(buf) as zf:
        for entry in zf.infolist():
            path = _fix_zip_filename(entry)
            if strip_prefix and not path.startswith(strip_prefix):
                continue
            rel_path = path[len(strip_prefix):]
            if not rel_path:
                continue

            if entry.is_dir():
                ensure_remote_dir(f"{group_dir}{rel_path}")
            else:
                if "/" in rel_path:
                    parent = rel_path.rsplit("/", 1)[0] + "/"
                    ensure_remote_dir(f"{group_dir}{parent}")
                data = io.BytesIO(zf.read(entry.filename))
                entry_remote = f"{group_dir}{rel_path}"
                upload_bytes(data, entry_remote)
                uploaded.append(entry_remote)
                if on_progress:
                    on_progress(len(uploaded), total)

    return uploaded


def upload_single_member_zip(buf: io.BytesIO, group_number: str, folder_name: str,
                             on_progress=None) -> list[str]:
    """Upload a single student zip into {UPLOAD_DIR}/{GROUP}/{folder_name}/.

    The zip contents are placed directly into the member folder
    (top-level entries from the archive, no nesting by archive structure).

    on_progress(uploaded_count, total_count) is called after each file upload.

    Returns list of uploaded remote paths.
    """
    upload_dir = get_webdav_upload_dir()
    ensure_remote_dir(upload_dir)

    group_dir = f"{upload_dir}{group_number}/"
    ensure_remote_dir(group_dir)

    member_dir = f"{group_dir}{folder_name}/"
    ensure_remote_dir(member_dir)

    total = _count_zip_files(buf)
    uploaded = []

    buf.seek(0)
    with zipfile.ZipFile(buf) as zf:
        for entry in zf.infolist():
            fixed_name = _fix_zip_filename(entry)
            if entry.is_dir():
                entry_dir = f"{member_dir}{fixed_name}"
                ensure_remote_dir(entry_dir)
            else:
                if "/" in fixed_name:
                    parent = fixed_name.rsplit("/", 1)[0] + "/"
                    ensure_remote_dir(f"{member_dir}{parent}")
                data = io.BytesIO(zf.read(entry.filename))
                entry_remote = f"{member_dir}{fixed_name}"
                upload_bytes(data, entry_remote)
                uploaded.append(entry_remote)
                if on_progress:
                    on_progress(len(uploaded), total)

    return uploaded


# Column header -> keyword to search for in student's filenames
REQUIRED_DOCUMENTS = {
    "Гарантийное письмо": "гарантийное письмо",
    "Заявление": "заявление",
    "Краткосрочный договор": "краткосрочный договор",
    "Характеристика": "характеристика",
}


def list_student_folders(group: str) -> list[str]:
    """List student folder names inside /Практики/{group}/ on Yandex Disk."""
    client = _get_client()
    upload_dir = get_webdav_upload_dir()
    group_dir = f"{upload_dir}{group}/"
    if not client.check(group_dir):
        return []
    items = client.list(group_dir)
    folders = []
    for item in items:
        name = item.strip("/")
        # Skip the directory itself (list() includes it) and any files
        if name and name != group:
            folders.append(name)
    return folders


def list_student_files(group: str, folder_name: str) -> set[str]:
    """List filenames inside a student's folder on Yandex Disk."""
    client = _get_client()
    upload_dir = get_webdav_upload_dir()
    student_dir = f"{upload_dir}{group}/{folder_name}/"
    if not client.check(student_dir):
        return set()
    items = client.list(student_dir)
    files = set()
    for item in items:
        name = item.strip("/")
        if name and name != folder_name:
            files.add(name)
    return files


def list_group_xlsx_files() -> list[str]:
    """List group xlsx files in the upload directory.

    Returns group names (without .xlsx extension).
    """
    client = _get_client()
    upload_dir = get_webdav_upload_dir()
    items = client.list(upload_dir)
    groups = []
    for item in items:
        name = item.strip("/")
        if re.match(r".+\.xlsx$", name, re.IGNORECASE):
            base = name[:-5]  # strip .xlsx
            if base.lower() == "администраторы":
                continue
            groups.append(base)
    return sorted(groups)


def download_xlsx(group: str) -> io.BytesIO:
    """Download {group}.xlsx from the upload directory into a BytesIO buffer."""
    client = _get_client()
    upload_dir = get_webdav_upload_dir()
    remote_path = f"{upload_dir}{group}.xlsx"
    buf = io.BytesIO()
    client.download_from(buf, remote_path)
    buf.seek(0)
    return buf


def sync_group_xlsx(group: str, on_progress=None) -> int:
    """Sync {group}.xlsx with actual files on Yandex Disk.

    For each student folder, checks which required documents are present
    and updates the corresponding xlsx columns.

    on_progress(current, total, student_name) is called for each student.

    Returns the number of cells updated.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    upload_dir = get_webdav_upload_dir()
    remote_path = f"{upload_dir}{group}.xlsx"
    client = _get_client()

    if not client.check(remote_path):
        return 0

    buf = io.BytesIO()
    client.download_from(buf, remote_path)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Map column headers to indices
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column - 1  # 0-based

    folders = list_student_folders(group)
    log.info("Sync %s: found %d folder(s) on disk", group, len(folders))
    # Build lookup: normalized name tuple -> set of files
    folder_files = {}
    for folder in folders:
        parts = folder.split("_")
        key = tuple(p for p in parts if p)
        files = list_student_files(group, folder)
        folder_files[key] = files
        log.info("  %s: %s", folder, ", ".join(sorted(files)) if files else "(пусто)")

    # Count students for progress
    student_rows = [r for r in ws.iter_rows(min_row=2) if r[0].value is not None]
    total = len(student_rows)

    updated = 0
    for i, row in enumerate(student_rows):
        surname = str(row[1].value or "").strip()
        name = str(row[2].value or "").strip()
        patronymic = str(row[3].value or "").strip()
        full_name = " ".join(filter(None, [surname, name, patronymic]))
        key = tuple(p for p in (surname, name, patronymic) if p)

        files = folder_files.get(key, set())
        log.info("Sync [%d/%d] %s — файлов на диске: %d", i + 1, total, full_name, len(files))

        row_changes = 0
        for col_header, keyword in REQUIRED_DOCUMENTS.items():
            col_idx = header_map.get(col_header)
            if col_idx is None:
                continue
            cell = row[col_idx]
            cell_val = str(cell.value or "").strip().lower()
            # Never touch cells marked as "не нужно" / "не надо"
            if cell_val in ("не нужно", "не надо"):
                log.info("  %s: %s (пропуск)", col_header, cell.value)
                continue
            has_file = any(keyword in f.replace("_", " ").lower() for f in files)
            current = cell_val in ("да", "yes", "1", "true")

            if has_file and not current:
                cell.value = "Да"
                cell.fill = green
                updated += 1
                row_changes += 1
                log.info("  %s: Нет -> Да", col_header)
            elif not has_file and current:
                cell.value = "Нет"
                cell.fill = red
                updated += 1
                row_changes += 1
                log.info("  %s: Да -> Нет", col_header)

        if on_progress:
            on_progress(i + 1, total, full_name)

    log.info("Sync %s done: %d cell(s) updated", group, updated)

    if updated:
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        client.upload_to(out, remote_path)

    wb.close()
    return updated


def list_students(group: str) -> list[dict]:
    """Return list of students from {group}.xlsx with name fields."""
    from openpyxl import load_workbook

    buf = download_xlsx(group)
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        students.append({
            "surname": str(row[1] or "").strip(),
            "name": str(row[2] or "").strip(),
            "patronymic": str(row[3] or "").strip(),
        })
    wb.close()
    return students


def set_student_comment(group: str, surname: str, name: str, patronymic: str, comment: str) -> bool:
    """Set the Комментарий cell for a student in {group}.xlsx. Returns True if updated."""
    from openpyxl import load_workbook

    upload_dir = get_webdav_upload_dir()
    remote_path = f"{upload_dir}{group}.xlsx"
    client = _get_client()

    if not client.check(remote_path):
        return False

    buf = io.BytesIO()
    client.download_from(buf, remote_path)
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active

    # Find Комментарий column
    comment_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == "Комментарий":
            comment_col = cell.column - 1
            break

    if comment_col is None:
        wb.close()
        return False

    target = tuple(p for p in (surname, name, patronymic) if p)

    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        s = str(row[1].value or "").strip()
        n = str(row[2].value or "").strip()
        p = str(row[3].value or "").strip()
        key = tuple(x for x in (s, n, p) if x)
        if key == target:
            row[comment_col].value = comment
            out = io.BytesIO()
            wb.save(out)
            out.seek(0)
            client.upload_to(out, remote_path)
            wb.close()
            return True

    wb.close()
    return False
